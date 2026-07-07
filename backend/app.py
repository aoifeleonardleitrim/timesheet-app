
import os  # Add this at the top with other imports
from flask import Flask, jsonify, send_from_directory, request
from flask_cors import CORS
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from flask_sqlalchemy import SQLAlchemy
from flask_bcrypt import Bcrypt
from datetime import timedelta, datetime
import json

app = Flask(__name__)

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///timesheet.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key')
app.config['JWT_SECRET_KEY'] = os.environ.get('JWT_SECRET_KEY', 'jwt-secret-key')
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(days=1)

CORS(app)
db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
jwt = JWTManager(app)

# ============================================
# MODELS
# ============================================

class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    first_name = db.Column(db.String(50), nullable=False)
    last_name = db.Column(db.String(50), nullable=False)
    role = db.Column(db.String(20), nullable=False)
    employee_id = db.Column(db.String(20), unique=True)
    position = db.Column(db.String(100))
    is_active = db.Column(db.Boolean, default=True)
    company_id = db.Column(db.Integer)
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())
    
    def set_password(self, password):
        self.password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
    
    def check_password(self, password):
        return bcrypt.check_password_hash(self.password_hash, password)
    
    def to_dict(self):
        return {
            'id': self.id,
            'email': self.email,
            'firstName': self.first_name,
            'lastName': self.last_name,
            'role': self.role,
            'employeeId': self.employee_id,
            'position': self.position,
            'isActive': self.is_active,
            'companyId': self.company_id
        }

class Company(db.Model):
    __tablename__ = 'companies'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    address = db.Column(db.String(200))
    phone = db.Column(db.String(20))
    email = db.Column(db.String(120))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())
    
    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'address': self.address,
            'phone': self.phone,
            'email': self.email,
            'isActive': self.is_active,
            'createdAt': self.created_at.isoformat() if self.created_at else None
        }

class Site(db.Model):
    __tablename__ = 'sites'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    address = db.Column(db.String(200))
    company_id = db.Column(db.Integer)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())
    
    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'address': self.address,
            'companyId': self.company_id,
            'isActive': self.is_active,
            'createdAt': self.created_at.isoformat() if self.created_at else None
        }

class EmployeeCompany(db.Model):
    __tablename__ = 'employee_companies'
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer)
    company_id = db.Column(db.Integer)
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())
    
    def to_dict(self):
        return {
            'id': self.id,
            'employeeId': self.employee_id,
            'companyId': self.company_id
        }

class CompanyRequest(db.Model):
    __tablename__ = 'company_requests'
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer)
    company_name = db.Column(db.String(100), nullable=False)
    address = db.Column(db.String(200))
    phone = db.Column(db.String(20))
    email = db.Column(db.String(120))
    status = db.Column(db.String(20), default='pending')
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())
    reviewed_at = db.Column(db.DateTime)
    reviewed_by = db.Column(db.Integer)
    
    def to_dict(self):
        employee = User.query.get(self.employee_id)
        reviewer = User.query.get(self.reviewed_by)
        return {
            'id': self.id,
            'employeeId': self.employee_id,
            'employeeName': f"{employee.first_name} {employee.last_name}" if employee else 'Unknown',
            'companyName': self.company_name,
            'address': self.address,
            'phone': self.phone,
            'email': self.email,
            'status': self.status,
            'createdAt': self.created_at.isoformat() if self.created_at else None,
            'reviewedAt': self.reviewed_at.isoformat() if self.reviewed_at else None,
            'reviewedBy': reviewer.to_dict() if reviewer else None
        }

class SiteRequest(db.Model):
    __tablename__ = 'site_requests'
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer)
    company_id = db.Column(db.Integer)
    site_name = db.Column(db.String(100), nullable=False)
    address = db.Column(db.String(200))
    status = db.Column(db.String(20), default='pending')
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())
    reviewed_at = db.Column(db.DateTime)
    reviewed_by = db.Column(db.Integer)
    
    def to_dict(self):
        employee = User.query.get(self.employee_id)
        company = Company.query.get(self.company_id)
        reviewer = User.query.get(self.reviewed_by)
        return {
            'id': self.id,
            'employeeId': self.employee_id,
            'employeeName': f"{employee.first_name} {employee.last_name}" if employee else 'Unknown',
            'companyId': self.company_id,
            'companyName': company.name if company else 'Unknown',
            'siteName': self.site_name,
            'address': self.address,
            'status': self.status,
            'createdAt': self.created_at.isoformat() if self.created_at else None,
            'reviewedAt': self.reviewed_at.isoformat() if self.reviewed_at else None,
            'reviewedBy': reviewer.to_dict() if reviewer else None
        }

class Timesheet(db.Model):
    __tablename__ = 'timesheets'
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer)
    company_id = db.Column(db.Integer)
    site_id = db.Column(db.Integer)
    week_start = db.Column(db.String(20), nullable=False)
    week_end = db.Column(db.String(20), nullable=False)
    days_data = db.Column(db.Text, default='{}')
    status = db.Column(db.String(20), default='draft')
    total_hours = db.Column(db.Float, default=0)
    notes = db.Column(db.Text)
    submitted_at = db.Column(db.DateTime)
    manager_name = db.Column(db.String(100))
    manager_signature = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())
    
    def to_dict(self):
        employee = User.query.get(self.employee_id)
        company = Company.query.get(self.company_id)
        site = Site.query.get(self.site_id)
        return {
            'id': self.id,
            'employeeId': self.employee_id,
            'employee': employee.to_dict() if employee else None,
            'companyId': self.company_id,
            'company': company.to_dict() if company else None,
            'siteId': self.site_id,
            'site': site.to_dict() if site else None,
            'weekStart': self.week_start,
            'weekEnd': self.week_end,
            'days': json.loads(self.days_data) if self.days_data else {},
            'status': self.status,
            'totalHours': self.total_hours,
            'notes': self.notes,
            'submittedAt': self.submitted_at.isoformat() if self.submitted_at else None,
            'managerName': self.manager_name,
            'hasSignature': bool(self.manager_signature),
            'createdAt': self.created_at.isoformat() if self.created_at else None
        }

# ============================================
# CREATE TABLES AND DEMO DATA
# ============================================

with app.app_context():
    db.drop_all()
    db.create_all()
    print("Database recreated")
    
    # Create demo employer
    employer = User(
        email='employer@company.com',
        first_name='John',
        last_name='Manager',
        role='employer',
        employee_id='EMP000001',
        position='Operations Manager'
    )
    employer.set_password('Admin123!')
    db.session.add(employer)
    
    # Create demo employee
    employee = User(
        email='employee@company.com',
        first_name='Sarah',
        last_name='Worker',
        role='employee',
        employee_id='EMP000002',
        position='Site Supervisor'
    )
    employee.set_password('WelcomeEMP000002!')
    db.session.add(employee)
    db.session.commit()
    print("Users created")
    
    # Get IDs
    employer_id = User.query.filter_by(email='employer@company.com').first().id
    employee_id = User.query.filter_by(email='employee@company.com').first().id
    
    # Create default company
    company = Company(
        name='ABC Construction',
        address='123 Builder Street, Dublin',
        phone='01-555-0100',
        email='info@abcconstruction.ie',
        is_active=True
    )
    db.session.add(company)
    db.session.commit()
    print("Company created")
    
    company_id = company.id
    
    # Assign employee to company
    emp_company = EmployeeCompany(
        employee_id=employee_id,
        company_id=company_id
    )
    db.session.add(emp_company)
    db.session.commit()
    print("Employee assigned to company")
    
    # Create default sites
    sites = [
        Site(name='Dublin Office', address='123 Builder Street, Dublin', company_id=company_id),
        Site(name='Cork Project', address='45 Harbour Road, Cork', company_id=company_id),
        Site(name='Galway Site', address='78 Riverside, Galway', company_id=company_id)
    ]
    for site in sites:
        db.session.add(site)
    db.session.commit()
    print("Sites created")
    
    print("Database initialized successfully")

# ============================================
# AUTH ROUTES
# ============================================

@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.json
    user = User.query.filter_by(email=data.get('email')).first()
    if not user or not user.check_password(data.get('password')):
        return jsonify({'error': 'Invalid credentials'}), 401
    token = create_access_token(identity=user.id)
    return jsonify({'token': token, 'user': user.to_dict()}), 200

@app.route('/api/auth/me', methods=['GET'])
@jwt_required()
def me():
    user = User.query.get(get_jwt_identity())
    return jsonify(user.to_dict()), 200

@app.route('/api/auth/change-password', methods=['POST'])
@jwt_required()
def change_password():
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    data = request.json
    current_password = data.get('currentPassword')
    new_password = data.get('newPassword')
    confirm_password = data.get('confirmPassword')
    
    if not user.check_password(current_password):
        return jsonify({'error': 'Current password is incorrect'}), 401
    
    if len(new_password) < 8:
        return jsonify({'error': 'Password must be at least 8 characters'}), 400
    
    if new_password != confirm_password:
        return jsonify({'error': 'Passwords do not match'}), 400
    
    user.set_password(new_password)
    db.session.commit()
    
    return jsonify({'message': 'Password changed successfully'}), 200

# ============================================
# EMPLOYEE ROUTES
# ============================================

@app.route('/api/employees', methods=['GET'])
@jwt_required()
def get_employees():
    user = User.query.get(get_jwt_identity())
    if user.role != 'employer':
        return jsonify({'error': 'Access denied'}), 403
    employees = User.query.filter_by(role='employee', is_active=True).all()
    return jsonify([e.to_dict() for e in employees]), 200

@app.route('/api/employees', methods=['POST'])
@jwt_required()
def create_employee():
    user = User.query.get(get_jwt_identity())
    if user.role != 'employer':
        return jsonify({'error': 'Only employers can create employees'}), 403
    
    data = request.json
    count = User.query.filter_by(role='employee').count()
    employee_id = f"EMP{count + 1:06d}"
    
    employee = User(
        email=data['email'],
        first_name=data['firstName'],
        last_name=data['lastName'],
        role='employee',
        position=data.get('position', ''),
        employee_id=employee_id,
        company_id=data.get('companyId')
    )
    default_password = f"Welcome{employee_id}!"
    employee.set_password(default_password)
    db.session.add(employee)
    db.session.commit()
    
    return jsonify({
        'message': 'Employee created', 
        'user': employee.to_dict(), 
        'defaultPassword': default_password
    }), 201

@app.route('/api/employees/<int:employee_id>', methods=['DELETE'])
@jwt_required()
def delete_employee(employee_id):
    user = User.query.get(get_jwt_identity())
    if user.role != 'employer':
        return jsonify({'error': 'Only employers can delete employees'}), 403
    
    employee = User.query.get(employee_id)
    if employee:
        employee.is_active = False
        db.session.commit()
    return jsonify({'message': 'Deleted'}), 200

# ============================================
# COMPANY ROUTES
# ============================================

@app.route('/api/companies', methods=['GET'])
@jwt_required()
def get_companies():
    user = User.query.get(get_jwt_identity())
    
    if user.role == 'employer':
        companies = Company.query.filter_by(is_active=True).all()
    else:
        emp_companies = EmployeeCompany.query.filter_by(employee_id=user.id).all()
        company_ids = [ec.company_id for ec in emp_companies]
        companies = Company.query.filter(Company.id.in_(company_ids), Company.is_active == True).all()
    
    return jsonify([c.to_dict() for c in companies]), 200

@app.route('/api/companies', methods=['POST'])
@jwt_required()
def create_company():
    user = User.query.get(get_jwt_identity())
    if user.role != 'employer':
        return jsonify({'error': 'Only employers can create companies'}), 403
    
    data = request.json
    company = Company(
        name=data.get('name'),
        address=data.get('address', ''),
        phone=data.get('phone', ''),
        email=data.get('email', ''),
        is_active=True
    )
    db.session.add(company)
    db.session.commit()
    return jsonify({'message': 'Company created', 'company': company.to_dict()}), 201

@app.route('/api/companies/<int:company_id>', methods=['PUT'])
@jwt_required()
def update_company(company_id):
    user = User.query.get(get_jwt_identity())
    if user.role != 'employer':
        return jsonify({'error': 'Only employers can update companies'}), 403
    
    company = Company.query.get(company_id)
    if not company:
        return jsonify({'error': 'Company not found'}), 404
    
    data = request.json
    company.name = data.get('name', company.name)
    company.address = data.get('address', company.address)
    company.phone = data.get('phone', company.phone)
    company.email = data.get('email', company.email)
    company.is_active = data.get('isActive', company.is_active)
    db.session.commit()
    return jsonify({'message': 'Company updated', 'company': company.to_dict()}), 200

@app.route('/api/companies/<int:company_id>', methods=['DELETE'])
@jwt_required()
def delete_company(company_id):
    user = User.query.get(get_jwt_identity())
    if user.role != 'employer':
        return jsonify({'error': 'Only employers can delete companies'}), 403
    
    company = Company.query.get(company_id)
    if company:
        company.is_active = False
        db.session.commit()
    return jsonify({'message': 'Company deactivated'}), 200

@app.route('/api/companies/<int:company_id>/assign', methods=['POST'])
@jwt_required()
def assign_employee_to_company(company_id):
    user = User.query.get(get_jwt_identity())
    if user.role != 'employer':
        return jsonify({'error': 'Only employers can assign employees'}), 403
    
    data = request.json
    employee_id = data.get('employeeId')
    
    existing = EmployeeCompany.query.filter_by(
        employee_id=employee_id,
        company_id=company_id
    ).first()
    
    if existing:
        return jsonify({'message': 'Already assigned'}), 200
    
    assignment = EmployeeCompany(
        employee_id=employee_id,
        company_id=company_id
    )
    db.session.add(assignment)
    db.session.commit()
    return jsonify({'message': 'Employee assigned to company'}), 201

# ============================================
# SITE ROUTES
# ============================================

@app.route('/api/sites', methods=['GET'])
@jwt_required()
def get_sites():
    user = User.query.get(get_jwt_identity())
    
    if user.role == 'employer':
        sites = Site.query.filter_by(is_active=True).all()
    else:
        emp_companies = EmployeeCompany.query.filter_by(employee_id=user.id).all()
        company_ids = [ec.company_id for ec in emp_companies]
        sites = Site.query.filter(Site.company_id.in_(company_ids), Site.is_active == True).all()
    
    return jsonify([s.to_dict() for s in sites]), 200

@app.route('/api/sites', methods=['POST'])
@jwt_required()
def create_site():
    user = User.query.get(get_jwt_identity())
    if user.role != 'employer':
        return jsonify({'error': 'Only employers can create sites'}), 403
    
    data = request.json
    site = Site(
        name=data.get('name'),
        address=data.get('address', ''),
        company_id=data.get('companyId'),
        is_active=True
    )
    db.session.add(site)
    db.session.commit()
    return jsonify({'message': 'Site created', 'site': site.to_dict()}), 201

@app.route('/api/sites/<int:site_id>', methods=['PUT'])
@jwt_required()
def update_site(site_id):
    user = User.query.get(get_jwt_identity())
    if user.role != 'employer':
        return jsonify({'error': 'Only employers can update sites'}), 403
    
    site = Site.query.get(site_id)
    if not site:
        return jsonify({'error': 'Site not found'}), 404
    
    data = request.json
    site.name = data.get('name', site.name)
    site.address = data.get('address', site.address)
    site.company_id = data.get('companyId', site.company_id)
    site.is_active = data.get('isActive', site.is_active)
    db.session.commit()
    return jsonify({'message': 'Site updated', 'site': site.to_dict()}), 200

@app.route('/api/sites/<int:site_id>', methods=['DELETE'])
@jwt_required()
def delete_site(site_id):
    user = User.query.get(get_jwt_identity())
    if user.role != 'employer':
        return jsonify({'error': 'Only employers can delete sites'}), 403
    
    site = Site.query.get(site_id)
    if site:
        site.is_active = False
        db.session.commit()
    return jsonify({'message': 'Site deactivated'}), 200

# ============================================
# COMPANY REQUEST ROUTES
# ============================================

@app.route('/api/company-requests', methods=['GET'])
@jwt_required()
def get_company_requests():
    user = User.query.get(get_jwt_identity())
    
    if user.role == 'employer':
        requests = CompanyRequest.query.filter_by(status='pending').order_by(CompanyRequest.created_at.desc()).all()
    else:
        requests = CompanyRequest.query.filter_by(employee_id=user.id).order_by(CompanyRequest.created_at.desc()).all()
    
    return jsonify([r.to_dict() for r in requests]), 200

@app.route('/api/company-requests', methods=['POST'])
@jwt_required()
def create_company_request():
    user = User.query.get(get_jwt_identity())
    
    data = request.json
    request_obj = CompanyRequest(
        employee_id=user.id,
        company_name=data.get('companyName'),
        address=data.get('address', ''),
        phone=data.get('phone', ''),
        email=data.get('email', ''),
        status='pending'
    )
    db.session.add(request_obj)
    db.session.commit()
    return jsonify({'message': 'Company request submitted', 'request': request_obj.to_dict()}), 201

@app.route('/api/company-requests/<int:request_id>/approve', methods=['POST'])
@jwt_required()
def approve_company_request(request_id):
    user = User.query.get(get_jwt_identity())
    if user.role != 'employer':
        return jsonify({'error': 'Only employers can approve requests'}), 403
    
    request_obj = CompanyRequest.query.get(request_id)
    if not request_obj:
        return jsonify({'error': 'Request not found'}), 404
    
    company = Company(
        name=request_obj.company_name,
        address=request_obj.address,
        phone=request_obj.phone,
        email=request_obj.email,
        is_active=True
    )
    db.session.add(company)
    db.session.commit()
    
    assignment = EmployeeCompany(
        employee_id=request_obj.employee_id,
        company_id=company.id
    )
    db.session.add(assignment)
    
    request_obj.status = 'approved'
    request_obj.reviewed_at = datetime.utcnow()
    request_obj.reviewed_by = user.id
    db.session.commit()
    
    return jsonify({'message': 'Company request approved and company created', 'company': company.to_dict()}), 200

@app.route('/api/company-requests/<int:request_id>/reject', methods=['POST'])
@jwt_required()
def reject_company_request(request_id):
    user = User.query.get(get_jwt_identity())
    if user.role != 'employer':
        return jsonify({'error': 'Only employers can reject requests'}), 403
    
    request_obj = CompanyRequest.query.get(request_id)
    if not request_obj:
        return jsonify({'error': 'Request not found'}), 404
    
    request_obj.status = 'rejected'
    request_obj.reviewed_at = datetime.utcnow()
    request_obj.reviewed_by = user.id
    db.session.commit()
    
    return jsonify({'message': 'Company request rejected'}), 200

# ============================================
# SITE REQUEST ROUTES
# ============================================

@app.route('/api/site-requests', methods=['GET'])
@jwt_required()
def get_site_requests():
    user = User.query.get(get_jwt_identity())
    
    if user.role == 'employer':
        requests = SiteRequest.query.filter_by(status='pending').order_by(SiteRequest.created_at.desc()).all()
    else:
        requests = SiteRequest.query.filter_by(employee_id=user.id).order_by(SiteRequest.created_at.desc()).all()
    
    return jsonify([r.to_dict() for r in requests]), 200

@app.route('/api/site-requests', methods=['POST'])
@jwt_required()
def create_site_request():
    user = User.query.get(get_jwt_identity())
    
    data = request.json
    request_obj = SiteRequest(
        employee_id=user.id,
        company_id=data.get('companyId'),
        site_name=data.get('siteName'),
        address=data.get('address', ''),
        status='pending'
    )
    db.session.add(request_obj)
    db.session.commit()
    return jsonify({'message': 'Site request submitted', 'request': request_obj.to_dict()}), 201

@app.route('/api/site-requests/<int:request_id>/approve', methods=['POST'])
@jwt_required()
def approve_site_request(request_id):
    user = User.query.get(get_jwt_identity())
    if user.role != 'employer':
        return jsonify({'error': 'Only employers can approve requests'}), 403
    
    request_obj = SiteRequest.query.get(request_id)
    if not request_obj:
        return jsonify({'error': 'Request not found'}), 404
    
    site = Site(
        name=request_obj.site_name,
        address=request_obj.address,
        company_id=request_obj.company_id,
        is_active=True
    )
    db.session.add(site)
    
    request_obj.status = 'approved'
    request_obj.reviewed_at = datetime.utcnow()
    request_obj.reviewed_by = user.id
    db.session.commit()
    
    return jsonify({'message': 'Site request approved and site created', 'site': site.to_dict()}), 200

@app.route('/api/site-requests/<int:request_id>/reject', methods=['POST'])
@jwt_required()
def reject_site_request(request_id):
    user = User.query.get(get_jwt_identity())
    if user.role != 'employer':
        return jsonify({'error': 'Only employers can reject requests'}), 403
    
    request_obj = SiteRequest.query.get(request_id)
    if not request_obj:
        return jsonify({'error': 'Request not found'}), 404
    
    request_obj.status = 'rejected'
    request_obj.reviewed_at = datetime.utcnow()
    request_obj.reviewed_by = user.id
    db.session.commit()
    
    return jsonify({'message': 'Site request rejected'}), 200

# ============================================
# TIMESHEET ROUTES
# ============================================

@app.route('/api/timesheets', methods=['GET'])
@jwt_required()
def get_timesheets():
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    
    if user.role == 'employer':
        timesheets = Timesheet.query.order_by(Timesheet.created_at.desc()).all()
    else:
        timesheets = Timesheet.query.filter_by(employee_id=user_id).order_by(Timesheet.created_at.desc()).all()
    
    return jsonify([t.to_dict() for t in timesheets]), 200

@app.route('/api/timesheets', methods=['POST'])
@jwt_required()
def create_timesheet():
    user_id = get_jwt_identity()
    data = request.json
    
    timesheet = Timesheet(
        employee_id=user_id,
        company_id=data.get('companyId'),
        site_id=data.get('siteId'),
        week_start=data['weekStart'],
        week_end=data['weekEnd'],
        notes=data.get('notes', ''),
        manager_name=data.get('managerName', ''),
        manager_signature=data.get('managerSignature')
    )
    timesheet.days_data = json.dumps(data.get('days', {}))
    
    total = 0
    days = data.get('days', {})
    for day, day_data in days.items():
        if day_data.get('status') == 'present' and day_data.get('startTime') and day_data.get('endTime'):
            try:
                start = datetime.strptime(day_data['startTime'], '%H:%M')
                end = datetime.strptime(day_data['endTime'], '%H:%M')
                lunch = float(day_data.get('lunchBreak') or 0)
                hours = (end - start).seconds / 3600 - lunch
                if hours > 0:
                    total += hours
            except:
                pass
    timesheet.total_hours = round(total, 2)
    
    db.session.add(timesheet)
    db.session.commit()
    
    return jsonify({'message': 'Timesheet created', 'timesheet': timesheet.to_dict()}), 201

@app.route('/api/timesheets/<int:timesheet_id>/submit', methods=['POST'])
@jwt_required()
def submit_timesheet(timesheet_id):
    timesheet = Timesheet.query.get(timesheet_id)
    if timesheet:
        timesheet.status = 'submitted'
        timesheet.submitted_at = datetime.utcnow()
        db.session.commit()
    return jsonify({'message': 'Submitted'}), 200

@app.route('/api/timesheets/<int:timesheet_id>/approve', methods=['POST'])
@jwt_required()
def approve_timesheet(timesheet_id):
    user = User.query.get(get_jwt_identity())
    if user.role != 'employer':
        return jsonify({'error': 'Access denied'}), 403
    
    timesheet = Timesheet.query.get(timesheet_id)
    if timesheet:
        timesheet.status = 'approved'
        db.session.commit()
    return jsonify({'message': 'Approved'}), 200

@app.route('/api/timesheets/<int:timesheet_id>/reject', methods=['POST'])
@jwt_required()
def reject_timesheet(timesheet_id):
    user = User.query.get(get_jwt_identity())
    if user.role != 'employer':
        return jsonify({'error': 'Access denied'}), 403
    
    timesheet = Timesheet.query.get(timesheet_id)
    if timesheet:
        timesheet.status = 'rejected'
        db.session.commit()
    return jsonify({'message': 'Rejected'}), 200

# ============================================
# EXPORT
# ============================================

@app.route('/api/export/excel', methods=['GET'])
@jwt_required()
def export_excel():
    from openpyxl import Workbook
    from io import BytesIO
    from flask import send_file
    
    user = User.query.get(get_jwt_identity())
    if user.role != 'employer':
        return jsonify({'error': 'Access denied'}), 403
    
    timesheets = Timesheet.query.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheets"
    
    headers = ['Employee', 'Company', 'Site', 'Week Start', 'Week End', 'Status', 'Hours', 'Manager']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    for row, ts in enumerate(timesheets, 2):
        employee = User.query.get(ts.employee_id)
        company = Company.query.get(ts.company_id)
        site = Site.query.get(ts.site_id)
        ws.cell(row=row, column=1, value=f"{employee.first_name} {employee.last_name}" if employee else 'Unknown')
        ws.cell(row=row, column=2, value=company.name if company else '')
        ws.cell(row=row, column=3, value=site.name if site else '')
        ws.cell(row=row, column=4, value=ts.week_start)
        ws.cell(row=row, column=5, value=ts.week_end)
        ws.cell(row=row, column=6, value=ts.status)
        ws.cell(row=row, column=7, value=ts.total_hours)
        ws.cell(row=row, column=8, value=ts.manager_name)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='timesheets.xlsx')

# ============================================
# SERVE FRONTEND
# ============================================

@app.route('/')
def index():
    return send_from_directory('../frontend', 'login.html')

@app.route('/<path:path>')
def frontend(path):
    return send_from_directory('../frontend', path)

# ============================================
# RUN
# ============================================

if __name__ == '__main__':
    print("\n" + "="*60)
    print("TIMESHEET SYSTEM RUNNING")
    print("="*60)
    print("\nAccess: http://localhost:5001")
    print("\nLogin:")
    print("  Employer: employer@company.com / Admin123!")
    print("  Employee: employee@company.com / WelcomeEMP000002!")
    print("\n" + "="*60 + "\n")
    app.run(debug=True, host='0.0.0.0', port=5001)